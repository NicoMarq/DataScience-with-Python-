# -*- coding: utf-8 -*-
"""
Created on Sun May 23 19:12:31 2021

@author: Nico
"""

import openpyxl
Excelworkbook=openpyxl.load_workbook('C:/MuestraTarjetaX2019.xlsx')
Excelsheet=Excelworkbook.active
from openpyxl import Workbook

#CLIENTES BLACK --> GANAN MAS DE 55mil PESOS
#ME CONCENTRO EN LA COLUMNA DE ID E INGRESOS

#DEFINO VARIABLES DE LAS COLUMNAS
ingresos_minimos=500000
id_clientes=Excelsheet['A']
ingresos_clientes=Excelsheet['C']


#DEFINO VARIABLE PARA CREAR UNA LISTA
lista_clientes_black=[list]


#ABRO EL LOOPS PARA RECORRER
indice=0
for celda in ingresos_clientes:
    if indice>0:
        if ingresos_clientes[indice].value>=ingresos_minimos:
            lista_clientes_black.append(id_clientes[indice].value)
            
    indice=indice+1
            
print(lista_clientes_black)

#CREANDO EXCEL NUEVO Y PONIENDO LISTA EN EL MISMO
NuevoExcel=Workbook()
NuevoExcel_sheet=NuevoExcel.active

NuevoExcel_sheet['A1']="IDs de Clientes recomendados para Black card"

indicenuevo=0
for cliente in lista_clientes_black:
    cellref=NuevoExcel_sheet.cell(row=indicenuevo+1, column=5)
    cellref.value=str(cliente)
    indicenuevo=indicenuevo+1
    
NuevoExcel.save('C:/Users/Nico/Documents/Cursos/Clientesblack.xlsx')



